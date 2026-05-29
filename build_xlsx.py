"""Build game_data.xlsx with 4 sheets: 卡牌 / 敌人 / 召唤物 / 公式.

卡牌 sheet is generated from card_data_dump.json — a live snapshot of CARD_DATA
(one row per family×tier). To refresh it after editing cards in game.js:
  1. Open the game in the browser (preview server, port 8766).
  2. Paste DUMP_SNIPPET (below) into the console; it returns {familyCount, rowCount, rows}.
  3. Save that JSON as card_data_dump.json next to this script.
  4. Run: python build_xlsx.py
(CARD_DATA factory functions are evaluated JS, so the live game is the only reliable
source for per-tier cost/value/desc — a static parse of game.js cannot resolve them.)

敌人 / 召唤物 / 公式 sheets are still hand-maintained reference data below.
"""
import json
import os
from openpyxl import Workbook

# Console snippet that produces card_data_dump.json (see module docstring).
DUMP_SNIPPET = r"""
(() => {
  const mk = window.__mkCard, C = window.__cards;
  const order = ['bronze','silver','gold','diamond'];
  const rzh = {bronze:'铜', silver:'银', gold:'金', diamond:'钻'};
  const rows = [];
  for (const fid of Object.keys(C)) {
    const f = C[fid], nm = f.name || {}, tiers = f.tiers || {};
    for (const t of Object.keys(tiers).sort((a,b)=>order.indexOf(a)-order.indexOf(b))) {
      const card = mk(fid, t);
      const d = tiers[t].desc, dd = (typeof d === 'function') ? d(card) : d;
      const zh = (dd && dd.zh) || (typeof dd==='string'?dd:''), en = (dd && dd.en) || (typeof dd==='string'?dd:'');
      rows.push({familyId:fid, name_zh:(nm.zh||nm), name_en:(nm.en||nm), emoji:f.emoji||'',
        excludedFromShop:!!f.excludedFromShop, tier:t, rarity_zh:rzh[t]||t,
        cost:card.cost, value:card.value, desc_zh:zh, desc_en:en});
    }
  }
  return {familyCount:Object.keys(C).length, rowCount:rows.length, rows};
})()
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
font_arial = 'Arial'
bold_white = Font(name=font_arial, bold=True, color='FFFFFF')
bold = Font(name=font_arial, bold=True)
normal = Font(name=font_arial)
thin_border = Border(
    left=Side(style='thin', color='888888'),
    right=Side(style='thin', color='888888'),
    top=Side(style='thin', color='888888'),
    bottom=Side(style='thin', color='888888'),
)

HEADER_FILL = PatternFill('solid', start_color='2A3242')

def header_row(sheet, columns):
    for col, name in enumerate(columns, 1):
        c = sheet.cell(row=1, column=col)
        c.value = name
        c.font = bold_white
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border

def style_data(sheet, n_rows, n_cols):
    for r in range(2, n_rows + 2):
        for c in range(1, n_cols + 1):
            cell = sheet.cell(row=r, column=c)
            cell.font = normal
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = thin_border

# ============== 卡牌 sheet（由 card_data_dump.json 动态生成，每行 = 一个 family×tier）==============
cards_sheet = wb.active
cards_sheet.title = '卡牌'

_dump_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'card_data_dump.json')
with open(_dump_path, encoding='utf-8') as _f:
    _dump = json.load(_f)

_tier_order = {'bronze': 0, 'silver': 1, 'gold': 2, 'diamond': 3}
CARD_ROWS = sorted(
    _dump['rows'],
    key=lambda r: (r.get('familyId', ''), _tier_order.get(r.get('tier', ''), 9)),
)

CARD_COLUMNS = ['家族ID', '名称', '英文名', 'Emoji', '稀有度', '费用', '价值', '效果(中文)', '效果(英文)', '商店排除']
header_row(cards_sheet, CARD_COLUMNS)
for i, r in enumerate(CARD_ROWS, 2):
    vals = [
        r.get('familyId', ''),
        r.get('name_zh', ''),
        r.get('name_en', ''),
        r.get('emoji', ''),
        r.get('rarity_zh', ''),
        r.get('cost', ''),
        r.get('value', ''),
        r.get('desc_zh', ''),
        r.get('desc_en', ''),
        '是' if r.get('excludedFromShop') else '',
    ]
    for col, val in enumerate(vals, 1):
        cards_sheet.cell(row=i, column=col, value=val)
style_data(cards_sheet, len(CARD_ROWS), len(CARD_COLUMNS))
# 列宽
widths = [16, 12, 16, 8, 8, 6, 6, 50, 50, 10]
for col, w in enumerate(widths, 1):
    cards_sheet.column_dimensions[chr(64 + col)].width = w
cards_sheet.row_dimensions[1].height = 24

# ============== 敌人 sheet ==============
enemies_sheet = wb.create_sheet('敌人')
ENEMIES = [
    # (Tier, name, hp, attack, speed, value, xpReward, shape, color, intent_desc)
    ('Tier 1', '哥布林', 8, 1, 90, 3, 2, '圆', '#7a8a4a', '接触造成 1 伤害，自爆'),
    ('Tier 1', '弓箭手', 3, 0, 40, 5, 3, '三角', '#a08060', '2 回合后射 1 颗弹（2 伤）'),
    ('Tier 1', '飞行兵', 5, 1, 130, 4, 3, '三角', '#4adcd0', '飞行接触 1 伤'),
    ('Tier 1', '突击兵', 7, 2, 180, 5, 4, '圆', '#ff5050', '高速冲刺 2 伤撞击'),
    ('Tier 2', '狙击手', 10, 0, 12, 7, 6, '三角', '#604070', '3 回合后高伤射击（5 伤）'),
    ('Tier 2', '弹射射手', 9, 0, 45, 6, 5, '圆', '#80d0d0', '2 回合后射弹射弹（弹 3）'),
    ('Tier 2', '追踪兵', 9, 0, 45, 6, 5, '圆', '#80a0d0', '2 回合后射追踪弹'),
    ('Tier 2', '治疗师', 10, 0, 40, 6, 5, '圆', '#60c060', '2 回合后治疗最近敌人 +3 HP'),
    ('Tier 2', '自爆球', 4, 6, 100, 5, 4, '圆', '#ffa040', '3 回合后自爆 6 伤 AOE'),
    ('Tier 2', '弹幕兵', 9, 0, 45, 7, 6, '圆', '#8080c0', '3 回合后 3 颗扇形弹（1 伤×3）'),
    ('Tier 3', '召唤师', 12, 0, 0, 9, 8, '方', '#702070', '3 回合后召唤 1 哥布林'),
    ('Tier 3', '指挥官', 11, 1, 45, 7, 7, '方', '#c08000', '2 回合后友军 +1 攻击'),
    ('Tier 3', '重甲兵', 25, 3, 35, 12, 9, '圆', '#444444', '接触造成 3 伤'),
    ('Tier 3', '法师', 8, 0, 28, 8, 7, '圆', '#a05ec0', '3 回合后 AOE 法术 3 伤'),
    ('Tier 3', '狂战士', 14, 2, 100, 9, 7, '圆', '#d04040', '1 回合后 +1 攻击（叠加），接触造成伤害'),
    ('Tier 3', '分裂者', 12, 0, 55, 6, 6, '圆', '#80c080', '接触 / 死亡分裂 2 个小型'),
    ('精英', '慢虫', 25, 1, 15, 8, 6, '方', '#a08040', '缓慢推进接触'),
    ('精英', '尖叫者', 7, 0, 40, 5, 5, '圆', '#c0a040', '3 回合后全敌人 +3 最大 HP'),
    ('精英', '时空法师', 11, 0, 45, 6, 6, '圆', '#6080c0', '2 回合后抽走玩家 1 法力'),
    ('Boss', '深渊魔王', 55, 4, 35, 25, 20, '圆', '#400040', '射 3 伤弹 / 召唤哥布林 / AOE 5 伤'),
    ('奖励', '金球', 999, 0, 0, 0, 0, '圆', '#ffd84a', '奖励目标。击中获得金币（指数递减）'),
]

header_row(enemies_sheet, ['等级', '名称', 'HP', '攻击', '速度', '价值费用', '经验奖励', '形状', '颜色', '行为描述'])
for i, row in enumerate(ENEMIES, 2):
    for col, val in enumerate(row, 1):
        enemies_sheet.cell(row=i, column=col, value=val)
# 敌人统一样式（无 tier 着色，等级作为文字列）
style_data(enemies_sheet, len(ENEMIES), 10)
widths_e = [8, 14, 6, 6, 6, 9, 9, 8, 12, 40]
for col, w in enumerate(widths_e, 1):
    enemies_sheet.column_dimensions[chr(64 + col)].width = w
enemies_sheet.row_dimensions[1].height = 24

# ============== 召唤物 sheet ==============
summons_sheet = wb.create_sheet('召唤物')
SUMMONS = [
    # (name, hp, attack, speed, bulletAttack, cooldown, decayRate, 描述)
    ('炮台', 3, 0, 0, 2, 1, 1, '固定炮台。每敌方回合射 1 颗子弹（继承玩家 buff）'),
    ('士兵', 5, 2, 60, '-', '-', 1, '近战推进。接触敌人造成 2 伤'),
    ('狙击塔', 6, 0, 0, 5, 2, 1, '高伤狙击。每 2 敌方回合射 5 伤弹'),
    ('护盾兵', 8, 2, 18, '-', '-', 1, '肉盾。缓慢推进 + 接触造成 2 伤'),
    ('无人机', 3, 0, '飞行', 1, 1, 1, '飞行单位。每敌方回合射 1 颗子弹'),
    ('奥弹炮台', 3, 0, 0, 2, 1, 1, '奥弹炮台。每敌方回合射 1 颗追踪奥弹'),
    ('弹射炮台', 3, 0, 0, 1, 1, 1, '炮台。子弹自带弹射 +3'),
    ('重型炮台', 7, 0, 0, 4, 1, 1, '高耐久炮台。每回合射 4 伤弹（战术撤退衍生）'),
    ('工兵', 3, 1, 95, '-', '-', 0, '快速近战。移速快、接触造成 1 伤、不衰减（双面间谍使用衍生）'),
    ('弓箭手', 4, 0, '-', 2, 1, 1, '远程弓兵。每敌方回合射 1 颗 2 伤箭（双面间谍弃置衍生）'),
]

header_row(summons_sheet, ['名称', 'HP', '攻击', '速度', '子弹伤害', '冷却(回合)', '回合衰减', '行为描述'])
for i, row in enumerate(SUMMONS, 2):
    for col, val in enumerate(row, 1):
        summons_sheet.cell(row=i, column=col, value=val)
style_data(summons_sheet, len(SUMMONS), 8)
widths_s = [14, 6, 6, 8, 10, 12, 10, 50]
for col, w in enumerate(widths_s, 1):
    summons_sheet.column_dimensions[chr(64 + col)].width = w
summons_sheet.row_dimensions[1].height = 24

# ============== 公式 sheet ==============
formulas_sheet = wb.create_sheet('公式')

def section(sheet, row, title):
    cell = sheet.cell(row=row, column=1, value=title)
    cell.font = Font(name=font_arial, bold=True, color='FFFFFF', size=12)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='left', vertical='center')
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    return row + 1

r = 1
# —— 1. 波次价值公式 ——
r = section(formulas_sheet, r, '1. 波次价值公式（每波 spawn 的"总价值"）')
formulas_sheet.cell(row=r, column=1, value='公式').font = bold
formulas_sheet.cell(row=r, column=2, value='value(w) = floor(8 + 3w + 0.3w²)').font = normal
formulas_sheet.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
r += 1
formulas_sheet.cell(row=r, column=1, value='波次 w').font = bold
formulas_sheet.cell(row=r, column=2, value='价值 (公式)').font = bold
formulas_sheet.cell(row=r, column=3, value='价值 (实际)').font = bold
formulas_sheet.cell(row=r, column=4, value='说明').font = bold
r += 1
for w_idx in range(11):
    formulas_sheet.cell(row=r, column=1, value=w_idx).font = normal
    formulas_sheet.cell(row=r, column=2, value=f'8 + 3×{w_idx} + 0.3×{w_idx}²').font = normal
    formulas_sheet.cell(row=r, column=3, value=int(8 + 3 * w_idx + 0.3 * w_idx * w_idx)).font = normal
    if w_idx == 0:
        formulas_sheet.cell(row=r, column=4, value='第 1 波（开战即生成）').font = normal
    r += 1
r += 1

# —— 2. 波次间隔（按 shopLevel） ——
r = section(formulas_sheet, r, '2. 波次间隔（按 shopLevel 决定每多少回合 spawn 一波）')
formulas_sheet.cell(row=r, column=1, value='shopLevel').font = bold
formulas_sheet.cell(row=r, column=2, value='间隔 m (回合)').font = bold
formulas_sheet.cell(row=r, column=3, value='候选数').font = bold
formulas_sheet.cell(row=r, column=4, value='说明').font = bold
r += 1
intervals = [5, 5, 4, 4, 3, 3, 3, 2]
for lv in range(1, 9):
    formulas_sheet.cell(row=r, column=1, value=lv).font = normal
    formulas_sheet.cell(row=r, column=2, value=intervals[lv - 1]).font = normal
    formulas_sheet.cell(row=r, column=3, value=min(8, 2 + lv)).font = normal
    if lv == 1: formulas_sheet.cell(row=r, column=4, value='初始等级').font = normal
    if lv == 8: formulas_sheet.cell(row=r, column=4, value='最高等级').font = normal
    r += 1
r += 1

# —— 3. XP 升级曲线 ——
r = section(formulas_sheet, r, '3. XP 升级曲线（min(100, floor(6 + 4(L-1) + 0.6(L-1)²))）')
formulas_sheet.cell(row=r, column=1, value='等级').font = bold
formulas_sheet.cell(row=r, column=2, value='升下一级所需 XP').font = bold
formulas_sheet.cell(row=r, column=3, value='累计 XP').font = bold
formulas_sheet.cell(row=r, column=4, value='说明').font = bold
r += 1
total_xp = 0
for lv in range(1, 16):
    need = min(100, int(6 + 4 * (lv - 1) + 0.6 * (lv - 1) * (lv - 1)))
    total_xp += need
    formulas_sheet.cell(row=r, column=1, value=lv).font = normal
    formulas_sheet.cell(row=r, column=2, value=need).font = normal
    formulas_sheet.cell(row=r, column=3, value=total_xp).font = normal
    if need == 100: formulas_sheet.cell(row=r, column=4, value='封顶').font = normal
    r += 1
r += 1

# —— 4. 敌人 HP 缩放 ——
r = section(formulas_sheet, r, '4. 敌人 HP 缩放（基础 HP × (1 + (玩家等级 - 1) × 0.08)）')
formulas_sheet.cell(row=r, column=1, value='玩家等级').font = bold
formulas_sheet.cell(row=r, column=2, value='HP 倍率').font = bold
formulas_sheet.cell(row=r, column=3, value='哥布林 HP (基础 8)').font = bold
formulas_sheet.cell(row=r, column=4, value='Boss HP (基础 55)').font = bold
r += 1
for lv in [1, 3, 5, 7, 10, 15, 20]:
    scale = 1 + (lv - 1) * 0.08
    formulas_sheet.cell(row=r, column=1, value=lv).font = normal
    formulas_sheet.cell(row=r, column=2, value=f'×{scale:.2f}').font = normal
    formulas_sheet.cell(row=r, column=3, value=int(8 * scale)).font = normal
    formulas_sheet.cell(row=r, column=4, value=int(55 * scale)).font = normal
    r += 1
r += 1

# —— 5. 金球金币奖励（指数递减） ——
r = section(formulas_sheet, r, '5. 金球金币奖励（gold = max(1, floor(damage / (5 + 累计伤害 × 0.1)))）')
formulas_sheet.cell(row=r, column=1, value='累计伤害').font = bold
formulas_sheet.cell(row=r, column=2, value='5 dmg → 金币').font = bold
formulas_sheet.cell(row=r, column=3, value='10 dmg → 金币').font = bold
formulas_sheet.cell(row=r, column=4, value='说明').font = bold
r += 1
for total in [0, 10, 20, 50, 100, 200]:
    g5 = max(1, int(5 / (5 + total * 0.1)))
    g10 = max(1, int(10 / (5 + total * 0.1)))
    formulas_sheet.cell(row=r, column=1, value=total).font = normal
    formulas_sheet.cell(row=r, column=2, value=g5).font = normal
    formulas_sheet.cell(row=r, column=3, value=g10).font = normal
    if total == 0: formulas_sheet.cell(row=r, column=4, value='初始：5 dmg=1 金').font = normal
    r += 1
r += 1

# —— 6. 刷新成本 + 商店升级 ——
r = section(formulas_sheet, r, '6. 商店成本（刷新随回合线性增长；升级按阈值）')
formulas_sheet.cell(row=r, column=1, value='项目').font = bold
formulas_sheet.cell(row=r, column=2, value='公式 / 表').font = bold
formulas_sheet.cell(row=r, column=3, value='示例').font = bold
formulas_sheet.cell(row=r, column=4, value='说明').font = bold
r += 1
formulas_sheet.cell(row=r, column=1, value='刷新成本').font = normal
formulas_sheet.cell(row=r, column=2, value='cost = 1 + refreshCount').font = normal
formulas_sheet.cell(row=r, column=3, value='1, 2, 3, 4, 5...').font = normal
formulas_sheet.cell(row=r, column=4, value='每次 Loot 重置 refreshCount=0').font = normal
r += 1
formulas_sheet.cell(row=r, column=1, value='额外购买').font = normal
formulas_sheet.cell(row=r, column=2, value='cost = 3 + 2 × (n - 1)').font = normal
formulas_sheet.cell(row=r, column=3, value='3, 5, 7, 9...').font = normal
formulas_sheet.cell(row=r, column=4, value='第 2 张起付费').font = normal
r += 1
formulas_sheet.cell(row=r, column=1, value='商店升级').font = normal
formulas_sheet.cell(row=r, column=2, value='SHOP_THRESHOLDS').font = normal
formulas_sheet.cell(row=r, column=3, value='Lv→Lv+1: 1/2/4/6/8/11/15 金').font = normal
formulas_sheet.cell(row=r, column=4, value='Lv 1→2 = 1 金；Lv 7→8 = 15 金').font = normal
r += 1
r += 1

# —— 7. 稀有度概率（按 shopLevel） ——
r = section(formulas_sheet, r, '7. 稀有度概率（按 shopLevel；RARITY_PROB）')
formulas_sheet.cell(row=r, column=1, value='shopLevel').font = bold
formulas_sheet.cell(row=r, column=2, value='⬜ Common').font = bold
formulas_sheet.cell(row=r, column=3, value='🟦 Rare').font = bold
formulas_sheet.cell(row=r, column=4, value='🟪 Epic / 🟧 Legendary').font = bold
r += 1
RARITY = [
    (1, '100%', '0%', '0% / 0%'),
    (2, '70%',  '30%','0% / 0%'),
    (3, '55%',  '35%','10% / 0%'),
    (4, '45%',  '33%','20% / 2%'),
    (5, '35%',  '35%','25% / 5%'),
    (6, '25%',  '35%','30% / 10%'),
    (7, '20%',  '30%','35% / 15%'),
    (8, '10%',  '30%','40% / 20%'),
]
for lv, c, ra, ep in RARITY:
    formulas_sheet.cell(row=r, column=1, value=lv).font = normal
    formulas_sheet.cell(row=r, column=2, value=c).font = normal
    formulas_sheet.cell(row=r, column=3, value=ra).font = normal
    formulas_sheet.cell(row=r, column=4, value=ep).font = normal
    r += 1
r += 1

# —— 8. 卡牌价值模型 ——
r = section(formulas_sheet, r, '8. 卡牌价值模型（按费用应提供的总价值）')
formulas_sheet.cell(row=r, column=1, value='费用').font = bold
formulas_sheet.cell(row=r, column=2, value='价值').font = bold
formulas_sheet.cell(row=r, column=3, value='效果').font = bold
formulas_sheet.cell(row=r, column=4, value='价值').font = bold
r += 1
COST_VALUE = [
    (0, 0, '子弹+1', 6),
    (1, 8, '子弹波数+1', 12),
    (2, 14, '弹射+1', 2),
    (3, 18, '穿透+1', 4),
]
for cost, v, eff, ev in COST_VALUE:
    formulas_sheet.cell(row=r, column=1, value=cost).font = normal
    formulas_sheet.cell(row=r, column=2, value=v).font = normal
    formulas_sheet.cell(row=r, column=3, value=eff).font = normal
    formulas_sheet.cell(row=r, column=4, value=ev).font = normal
    r += 1
r += 1
MORE_EFFECTS = [
    ('追踪', 4, '展露', '基础效果价值 × 3'),
    ('火焰+1', '约 6', '洗入', '每张 -7'),
    ('连击需求', '-6 -(0.5×额外要求)', '弃置触发', '-7'),
    ('召唤物', '8 ~ 25', '随机', '效果平均 -1'),
]
for a, b, c, d in MORE_EFFECTS:
    formulas_sheet.cell(row=r, column=1, value=a).font = normal
    formulas_sheet.cell(row=r, column=2, value=b).font = normal
    formulas_sheet.cell(row=r, column=3, value=c).font = normal
    formulas_sheet.cell(row=r, column=4, value=d).font = normal
    r += 1

# 公式 sheet 整体样式
formulas_sheet.column_dimensions['A'].width = 22
formulas_sheet.column_dimensions['B'].width = 30
formulas_sheet.column_dimensions['C'].width = 30
formulas_sheet.column_dimensions['D'].width = 40
for row_idx in range(1, r):
    for c in range(1, 5):
        cell = formulas_sheet.cell(row=row_idx, column=c)
        if cell.value and not cell.font.bold:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = thin_border

# 冻结首行
for sh in (cards_sheet, enemies_sheet, summons_sheet):
    sh.freeze_panes = 'A2'

wb.save('game_data.xlsx')
print('Saved game_data.xlsx')
print('Cards:', len(CARD_ROWS), '(families:', _dump['familyCount'], ') Enemies:', len(ENEMIES), 'Summons:', len(SUMMONS))
